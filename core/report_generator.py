"""Generate Excel audit report from invoices and flags."""
import io
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


def generate(invoices: list, flags: list) -> bytes:
    buffer = io.BytesIO()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        # Sheet 1: Invoices
        inv_df = pd.DataFrame(invoices)
        if not inv_df.empty:
            inv_df = inv_df[
                [c for c in ["id", "filename", "vendor_name", "invoice_number", "invoice_date", "grand_total", "extraction_method", "extraction_confidence", "processed_at"] if c in inv_df.columns]
            ]
        inv_df.to_excel(writer, sheet_name="Invoices", index=False)

        # Sheet 2: Flags
        flags_df = pd.DataFrame(flags)
        if not flags_df.empty:
            flags_df = flags_df[
                [c for c in ["id", "invoice_id", "flag_type", "description", "billed_amount", "correct_amount", "overcharge", "severity", "status", "created_at"] if c in flags_df.columns]
            ]
        flags_df.to_excel(writer, sheet_name="Flags", index=False)

    buffer.seek(0)
    wb = load_workbook(buffer)
    for sheet in wb.worksheets:
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()
